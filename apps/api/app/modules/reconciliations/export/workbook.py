"""Builder do workbook Excel — 5 abas (S14 BACK 10.1).

Função pura: recebe `ExportPayload` (DTO já normalizado pelo service) e
devolve `BytesIO` com o XLSX em memória.

Princípios:
    - Zero I/O — nada de arquivo em disco; tudo em `BytesIO`.
    - Zero crypto / DB — o service entrega tudo descriptografado.
    - Sem float — valores monetários sempre `Decimal`.
    - Sem strings mágicas — labels (cabeçalhos, status legível) ficam em
      constants locais para facilitar i18n no futuro.

Estrutura por aba:
    1. Resumo            — bloco-livre (cabeçalho + saldos + indicadores
                           + anomalias + rodapé).
    2. Movimentação x    — tabela com cor por situation.
       Lançamento
    3. Divergências Omie — tabela com cor vermelha para "Atrasado".
    4. Sem Omie          — tabela amarela com `final_situation = sem_omie`.
    5. Anomalias         — tabela ordenada por severity (crítica primeiro).
"""

from __future__ import annotations

from collections.abc import Sequence
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook

from app.modules.reconciliations.export.schemas import (
    AnomalyRow,
    ExportPayload,
    FileEntryRow,
    OmieDivergenceRow,
    SemOmieRow,
    SummarySheetData,
)
from app.modules.reconciliations.export.styles import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    BORDER_CELL,
    FILL_ATRASADO,
    FILL_CRITICAL_UNRESOLVED,
    FILL_HEADER,
    FILL_SEM_OMIE,
    FONT_BOLD,
    FONT_CRITICAL_UNRESOLVED,
    FONT_DEFAULT,
    FONT_FOOTER,
    FONT_HEADER,
    FONT_SUBTITLE,
    FONT_TITLE,
    NUMBER_FORMAT_BRL,
    NUMBER_FORMAT_DATE,
    fill_for_situation,
)

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


# ----------------------------------------------------------------------
# Constants — labels (apenas pt-BR no relatório final)
# ----------------------------------------------------------------------

SHEET_NAME_SUMMARY = "Resumo"
SHEET_NAME_MOVIMENTACAO = "Movimentação x Lançamento"
SHEET_NAME_DIVERGENCIAS = "Divergências Omie"
SHEET_NAME_SEM_OMIE = "Sem Omie"
SHEET_NAME_ANOMALIAS = "Anomalias"

# Severity ordering (aba 5)
_SEVERITY_RANK = {"critical": 0, "moderate": 1, "info": 2}

# Tradução exibida no Excel (aba 2 e 4 mostram raw situation pra rastreio
# do analista; aba 5 traduz severity e detected_by porque o leitor final
# é não-técnico).
_SEVERITY_LABEL = {
    "critical": "Crítica",
    "moderate": "Moderada",
    "info": "Informativa",
}

_DETECTED_BY_LABEL = {
    "ai": "IA",
    "manual": "Manual",
}

_OMIE_STATUS_LABEL = {
    "Conciliado": "Conciliado",
    "Atrasado": "Atrasado",
    "Previsto": "Previsto",
}

_SITUATION_LABEL = {
    "conciliado": "Conciliado",
    "sem_omie": "Sem Omie",
    "ignorado": "Ignorado",
}

# Placeholder usado quando dado opcional vem como None (ex: supplier do
# cache L2 expirou). "—" (em-dash) é mais legível que "N/D" pro analista.
_PLACEHOLDER = "—"


# ======================================================================
# Entry point
# ======================================================================


def build_workbook(payload: ExportPayload) -> BytesIO:
    """Gera o workbook completo a partir do payload normalizado.

    Returns:
        `BytesIO` posicionado em 0 — caller pode passar direto pro
        `StreamingResponse` sem precisar de `seek(0)`.
    """
    wb = Workbook()
    # `Workbook()` cria uma sheet default "Sheet"; reusamos como aba 1.
    summary_ws = wb.active
    if summary_ws is None:
        # Defensivo — openpyxl sempre cria, mas type checker não sabe.
        summary_ws = wb.create_sheet()
    summary_ws.title = SHEET_NAME_SUMMARY

    _build_sheet1_summary(summary_ws, payload.summary)
    _build_sheet2_movimentacao(
        wb.create_sheet(title=SHEET_NAME_MOVIMENTACAO),
        payload.file_entries,
    )
    _build_sheet3_divergencias(
        wb.create_sheet(title=SHEET_NAME_DIVERGENCIAS),
        payload.omie_divergences,
    )
    _build_sheet4_sem_omie(
        wb.create_sheet(title=SHEET_NAME_SEM_OMIE),
        payload.sem_omie,
    )
    _build_sheet5_anomalias(
        wb.create_sheet(title=SHEET_NAME_ANOMALIAS),
        payload.anomalies,
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ======================================================================
# Aba 1 — Resumo
# ======================================================================


def _build_sheet1_summary(ws: Worksheet, data: SummarySheetData) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    # Título
    ws["A1"] = "Relatório de Conciliação"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:D1")

    # Cabeçalho (3 linhas com label/valor)
    ws["A3"] = "Cliente"
    ws["A3"].font = FONT_BOLD
    ws["B3"] = data.client_name
    ws["A4"] = "Banco"
    ws["A4"].font = FONT_BOLD
    ws["B4"] = data.bank_name
    ws["A5"] = "Conta"
    ws["A5"].font = FONT_BOLD
    ws["B5"] = data.account_name
    ws["A6"] = "Período"
    ws["A6"].font = FONT_BOLD
    ws["B6"] = data.period_pt_br

    ws["A8"] = "Elaborado por Hologram Gestão"
    ws["A8"].font = FONT_SUBTITLE
    ws.merge_cells("A8:D8")

    # ---- Tabela de saldos ------------------------------------------------
    ws["A10"] = "Saldos do período"
    ws["A10"].font = FONT_SUBTITLE
    ws.merge_cells("A10:D10")

    headers = ["Saldo inicial", "Saldo final (arquivo)", "Saldo final (Omie)", "Status"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=label)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_CELL

    ws.cell(row=12, column=1, value=data.balance_start).number_format = NUMBER_FORMAT_BRL
    ws.cell(row=12, column=2, value=data.balance_end_file).number_format = NUMBER_FORMAT_BRL
    ws.cell(row=12, column=3, value=data.balance_end_omie).number_format = NUMBER_FORMAT_BRL

    status_label = _resolve_balance_status(data.balance_difference)
    status_cell = ws.cell(row=12, column=4, value=status_label)
    status_cell.alignment = ALIGN_CENTER
    if status_label == "Divergente":
        # Mesma paleta da Aba 3 — destaca divergência sem inventar cor nova.
        status_cell.fill = FILL_ATRASADO
        status_cell.font = FONT_BOLD

    for col in range(1, 5):
        ws.cell(row=12, column=col).border = BORDER_CELL

    # ---- Indicadores do período ------------------------------------------
    ws["A14"] = "Indicadores do período"
    ws["A14"].font = FONT_SUBTITLE
    ws.merge_cells("A14:D14")

    indicators = [
        ("Movimentações no extrato", data.total_file_entries),
        ("Conciliadas", data.conciliated_count),
        ("Sem correspondência Omie", data.sem_omie_count),
        ("Omie sem arquivo", data.omie_sem_arquivo_count),
        ("Ignoradas", data.ignored_count),
    ]
    for offset, (label, value) in enumerate(indicators):
        row = 15 + offset
        ws.cell(row=row, column=1, value=label).font = FONT_BOLD
        ws.cell(row=row, column=2, value=value)

    # ---- Anomalias -------------------------------------------------------
    anomalies_start = 15 + len(indicators) + 1
    ws.cell(row=anomalies_start, column=1, value="Anomalias").font = FONT_SUBTITLE
    ws.merge_cells(start_row=anomalies_start, start_column=1, end_row=anomalies_start, end_column=4)

    breakdown = [
        ("Total", data.anomaly_total),
        ("Críticas", data.anomaly_critical),
        ("Moderadas", data.anomaly_moderate),
        ("Informativas", data.anomaly_info),
        ("Resolvidas", data.anomaly_resolved),
        ("Críticas não resolvidas", data.anomaly_critical_unresolved),
    ]
    for offset, (label, value) in enumerate(breakdown, start=1):
        row = anomalies_start + offset
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = FONT_BOLD
        value_cell = ws.cell(row=row, column=2, value=value)
        # Destaque vermelho em "Críticas não resolvidas" > 0 (mesmo zerado
        # a linha existe — assim o leitor sabe que o KPI foi avaliado).
        if label == "Críticas não resolvidas" and value > 0:
            label_cell.fill = FILL_CRITICAL_UNRESOLVED
            label_cell.font = FONT_CRITICAL_UNRESOLVED
            value_cell.fill = FILL_CRITICAL_UNRESOLVED
            value_cell.font = FONT_CRITICAL_UNRESOLVED

    # ---- Rodapé ----------------------------------------------------------
    footer_row = anomalies_start + len(breakdown) + 2
    footer_cell = ws.cell(
        row=footer_row,
        column=1,
        value=(
            f"Gerado em {data.generated_at_brt.strftime('%d/%m/%Y %H:%M')} (BRT) "
            f"por {data.generated_by_email} via FaturIA"
        ),
    )
    footer_cell.font = FONT_FOOTER
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=4)


def _resolve_balance_status(balance_difference: Decimal | None) -> str:
    """Aba 1: Conferido se diferença ≤ R$ 0,01, Divergente caso contrário."""
    if balance_difference is None:
        return "Indisponível"
    if abs(balance_difference) <= Decimal("0.01"):
        return "Conferido"
    return "Divergente"


# ======================================================================
# Aba 2 — Movimentação x Lançamento
# ======================================================================


_SHEET2_HEADERS = [
    ("Data", 14),
    ("Descrição", 48),
    ("Valor", 16),
    ("Saldo", 16),
    ("Fornecedor Omie", 32),
    ("Categoria Omie", 28),
    ("Situação", 18),
    ("Observação", 36),
]


def _build_sheet2_movimentacao(ws: Worksheet, rows: Sequence[FileEntryRow]) -> None:
    _write_table_header(ws, _SHEET2_HEADERS)

    for offset, row in enumerate(rows):
        excel_row = offset + 2
        ws.cell(
            row=excel_row, column=1, value=row.transaction_date
        ).number_format = NUMBER_FORMAT_DATE
        ws.cell(row=excel_row, column=2, value=row.description).alignment = ALIGN_LEFT
        ws.cell(row=excel_row, column=3, value=row.amount).number_format = NUMBER_FORMAT_BRL
        balance_cell = ws.cell(row=excel_row, column=4, value=row.balance)
        balance_cell.number_format = NUMBER_FORMAT_BRL
        ws.cell(row=excel_row, column=5, value=row.supplier or _PLACEHOLDER)
        ws.cell(row=excel_row, column=6, value=row.category or _PLACEHOLDER)
        ws.cell(
            row=excel_row,
            column=7,
            value=_SITUATION_LABEL.get(row.situation, row.situation),
        ).alignment = ALIGN_CENTER
        ws.cell(row=excel_row, column=8, value=row.user_note or "").alignment = ALIGN_LEFT

        fill = fill_for_situation(row.situation)
        if fill is not None:
            for col_idx in range(1, len(_SHEET2_HEADERS) + 1):
                ws.cell(row=excel_row, column=col_idx).fill = fill

    ws.freeze_panes = "A2"


# ======================================================================
# Aba 3 — Divergências Omie
# ======================================================================


_SHEET3_HEADERS = [
    ("Data", 14),
    ("Fornecedor", 32),
    ("Categoria", 28),
    ("Valor", 16),
    ("Status Omie", 16),
    ("Observação", 36),
]


def _build_sheet3_divergencias(ws: Worksheet, rows: Sequence[OmieDivergenceRow]) -> None:
    _write_table_header(ws, _SHEET3_HEADERS)

    for offset, row in enumerate(rows):
        excel_row = offset + 2
        ws.cell(
            row=excel_row, column=1, value=row.transaction_date
        ).number_format = NUMBER_FORMAT_DATE
        ws.cell(row=excel_row, column=2, value=row.supplier or _PLACEHOLDER)
        ws.cell(row=excel_row, column=3, value=row.category or _PLACEHOLDER)
        amount_cell = ws.cell(row=excel_row, column=4, value=row.amount)
        amount_cell.number_format = NUMBER_FORMAT_BRL
        if row.amount is None:
            amount_cell.value = _PLACEHOLDER
            amount_cell.alignment = ALIGN_RIGHT
        ws.cell(
            row=excel_row,
            column=5,
            value=_OMIE_STATUS_LABEL.get(row.omie_status, row.omie_status),
        ).alignment = ALIGN_CENTER
        ws.cell(row=excel_row, column=6, value=row.user_note or "").alignment = ALIGN_LEFT

        if row.omie_status == "Atrasado":
            for col_idx in range(1, len(_SHEET3_HEADERS) + 1):
                ws.cell(row=excel_row, column=col_idx).fill = FILL_ATRASADO

    ws.freeze_panes = "A2"


# ======================================================================
# Aba 4 — Sem Omie
# ======================================================================


_SHEET4_HEADERS = [
    ("Data", 14),
    ("Descrição", 48),
    ("Valor", 16),
    ("Observação", 36),
]


def _build_sheet4_sem_omie(ws: Worksheet, rows: Sequence[SemOmieRow]) -> None:
    _write_table_header(ws, _SHEET4_HEADERS)

    for offset, row in enumerate(rows):
        excel_row = offset + 2
        ws.cell(
            row=excel_row, column=1, value=row.transaction_date
        ).number_format = NUMBER_FORMAT_DATE
        ws.cell(row=excel_row, column=2, value=row.description).alignment = ALIGN_LEFT
        ws.cell(row=excel_row, column=3, value=row.amount).number_format = NUMBER_FORMAT_BRL
        ws.cell(row=excel_row, column=4, value=row.user_note or "").alignment = ALIGN_LEFT

        for col_idx in range(1, len(_SHEET4_HEADERS) + 1):
            ws.cell(row=excel_row, column=col_idx).fill = FILL_SEM_OMIE

    ws.freeze_panes = "A2"


# ======================================================================
# Aba 5 — Anomalias
# ======================================================================


_SHEET5_HEADERS = [
    ("Severidade", 16),
    ("Tipo", 36),
    ("Linha relacionada", 28),
    ("Detectado por", 16),
    ("Status", 16),
    ("Nota de resolução", 48),
]


def _build_sheet5_anomalias(ws: Worksheet, rows: Sequence[AnomalyRow]) -> None:
    _write_table_header(ws, _SHEET5_HEADERS)

    ordered = sorted(
        rows,
        key=lambda r: (_SEVERITY_RANK.get(r.severity, 99), r.resolved),
    )

    for offset, row in enumerate(ordered):
        excel_row = offset + 2
        ws.cell(
            row=excel_row,
            column=1,
            value=_SEVERITY_LABEL.get(row.severity, row.severity),
        ).alignment = ALIGN_CENTER
        ws.cell(row=excel_row, column=2, value=row.type_name).alignment = ALIGN_LEFT
        ws.cell(row=excel_row, column=3, value=row.related_line).alignment = ALIGN_LEFT
        ws.cell(
            row=excel_row,
            column=4,
            value=_DETECTED_BY_LABEL.get(row.detected_by, row.detected_by),
        ).alignment = ALIGN_CENTER
        ws.cell(
            row=excel_row,
            column=5,
            value="Resolvida" if row.resolved else "Pendente",
        ).alignment = ALIGN_CENTER
        ws.cell(row=excel_row, column=6, value=row.resolution_note or "").alignment = ALIGN_LEFT

        # Destaque (somente texto bold) em críticas não resolvidas — não
        # pinta fundo da linha pra não competir com a aba 1 (que usa o
        # mesmo vermelho como cor de status).
        if row.severity == "critical" and not row.resolved:
            for col_idx in range(1, len(_SHEET5_HEADERS) + 1):
                ws.cell(row=excel_row, column=col_idx).font = FONT_CRITICAL_UNRESOLVED

    ws.freeze_panes = "A2"


# ======================================================================
# Helpers
# ======================================================================

_COLUMN_LETTERS = [
    "A",
    "B",
    "C",
    "D",
    "E",
    "F",
    "G",
    "H",
    "I",
    "J",
    "K",
    "L",
    "M",
    "N",
    "O",
    "P",
]


def _write_table_header(ws: Worksheet, headers: Sequence[tuple[str, int]]) -> None:
    """Escreve a linha 1 de cabeçalho + define a largura de cada coluna.

    headers: lista de `(label, width)`.
    """
    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_CELL
        letter = _COLUMN_LETTERS[col_idx - 1]
        ws.column_dimensions[letter].width = width

    # Default font para corpo da tabela — openpyxl não tem um "default por
    # planilha", só aplicamos em runtime quando criamos a célula. Como as
    # células vazias herdam o default da workbook (Calibri 11), basta
    # garantir que o cabeçalho destaca.
    _ = FONT_DEFAULT  # mantém o import vivo pra qualquer evolução

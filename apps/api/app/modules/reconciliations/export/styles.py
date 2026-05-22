"""Estilos reusáveis do relatório Excel (S14 BACK 10.1).

Paleta + fontes + borders + helpers que pintam linhas inteiras. Toda
formatação visual do workbook deve passar por aqui — assim, mudar a
identidade visual do relatório é uma edição em 1 arquivo.

Convenções:
    - Hex de cor SEM `#` (openpyxl exige ARGB sem prefixo).
    - `PatternFill` solid em tudo (sem gradient — Excel 2007+ não
      renderiza consistentemente).
    - Formato monetário BRL: `R$ #,##0.00` (number_format do openpyxl
      respeita locale do Excel, mas o caractere `R$` literal funciona em
      qualquer idioma de Excel).
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Paleta de cores
# ---------------------------------------------------------------------------

# Cabeçalho de tabela (cinza escuro)
COLOR_HEADER_BG = "1F2937"  # gray-800
COLOR_HEADER_FG = "FFFFFF"

# Aba 2 — Movimentação por situation (linhas inteiras)
COLOR_ROW_CONCILIADO = "E2F5E2"  # verde claro
COLOR_ROW_SEM_OMIE = "FFF4CC"  # amarelo claro
COLOR_ROW_IGNORADO = "EEEEEE"  # cinza claro

# Aba 3 — Atrasado em vermelho claro
COLOR_ROW_ATRASADO = "FFE2E2"

# Aba 1 — destaque para anomalias críticas não resolvidas
COLOR_CRITICAL_UNRESOLVED_BG = "FFB3B3"
COLOR_CRITICAL_UNRESOLVED_FG = "7F1D1D"

# Linha de divisão (mesma cor em todas as abas)
COLOR_BORDER = "D1D5DB"

# ---------------------------------------------------------------------------
# Number formats
# ---------------------------------------------------------------------------

NUMBER_FORMAT_BRL = "R$ #,##0.00;[Red]-R$ #,##0.00"
NUMBER_FORMAT_DATE = "DD/MM/YYYY"

# ---------------------------------------------------------------------------
# Fonts
# ---------------------------------------------------------------------------

FONT_DEFAULT = Font(name="Calibri", size=11)
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_TITLE = Font(name="Calibri", size=16, bold=True)
FONT_SUBTITLE = Font(name="Calibri", size=12, bold=True, color="374151")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER_FG)
FONT_FOOTER = Font(name="Calibri", size=9, italic=True, color="6B7280")
FONT_CRITICAL_UNRESOLVED = Font(
    name="Calibri", size=11, bold=True, color=COLOR_CRITICAL_UNRESOLVED_FG
)

# ---------------------------------------------------------------------------
# Fills
# ---------------------------------------------------------------------------


def _solid_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


FILL_HEADER = _solid_fill(COLOR_HEADER_BG)
FILL_CONCILIADO = _solid_fill(COLOR_ROW_CONCILIADO)
FILL_SEM_OMIE = _solid_fill(COLOR_ROW_SEM_OMIE)
FILL_IGNORADO = _solid_fill(COLOR_ROW_IGNORADO)
FILL_ATRASADO = _solid_fill(COLOR_ROW_ATRASADO)
FILL_CRITICAL_UNRESOLVED = _solid_fill(COLOR_CRITICAL_UNRESOLVED_BG)

# ---------------------------------------------------------------------------
# Borders / alignment
# ---------------------------------------------------------------------------

_thin_side = Side(style="thin", color=COLOR_BORDER)
BORDER_CELL = Border(top=_thin_side, bottom=_thin_side, left=_thin_side, right=_thin_side)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------------------------
# Mapping de situation → fill (aba 2)
# ---------------------------------------------------------------------------


def fill_for_situation(situation: str) -> PatternFill | None:
    """Retorna o `PatternFill` da situation. `None` quando não há regra."""
    if situation == "conciliado":
        return FILL_CONCILIADO
    if situation == "sem_omie":
        return FILL_SEM_OMIE
    if situation == "ignorado":
        return FILL_IGNORADO
    return None

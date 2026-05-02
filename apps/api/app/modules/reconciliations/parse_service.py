"""Service de parsing IA (S9 — BACK 7.1).

Responsabilidade: transformar bytes de upload em `ExtractedStatement`
chamando a Anthropic via `AnthropicClient`. Stateless — não persiste nada;
quem persiste é S10/BACK 8.1.

Fronteiras (CLAUDE.md §3.8 + §3.10):
    - Validação de tamanho, extensão e magic bytes acontece AQUI no servidor,
      mesmo que o front (S8) já tenha validado. Servidor é a fonte da
      verdade; front é UX.
    - Arquivo NUNCA toca disco. Bytes em memória, descartados ao final do
      request.

Pipeline para cada formato suportado:
    PDF   → bytes brutos   →   AnthropicClient (document base64).
    CSV   → decode utf-8   →   AnthropicClient (text block).
    XLSX  → openpyxl       →   render TSV   →   AnthropicClient (text block).
    XLS   → não suportado nesta versão (xlrd não está nas deps).
"""

from __future__ import annotations

from io import BytesIO
from pathlib import PurePosixPath
from typing import TYPE_CHECKING

import openpyxl

from app.core.exceptions import ValidationAppError
from app.core.logging import get_logger
from app.integrations.anthropic.schemas import ExtractedStatement
from app.utils.magic_bytes import FileType, validate_upload_type

if TYPE_CHECKING:
    from app.integrations.anthropic.client import AnthropicClient

log = get_logger(__name__)

# Conjunto aceito pelo endpoint de parse. Mantém XLS de fora — o checklist da
# UI lista XLS, mas para BACK 7.1 não temos um decoder confiável (xlrd<2.0
# não está nas deps). Se o cliente subir .xls, recusamos com mensagem clara.
_ALLOWED_FOR_PARSE: set[FileType] = {
    FileType.PDF,
    FileType.CSV,
    FileType.XLSX,
}

_DOCUMENT_KIND: dict[FileType, str] = {
    FileType.PDF: "extrato/fatura em PDF",
    FileType.CSV: "extrato/fatura em CSV",
    FileType.XLSX: "extrato/fatura em planilha XLSX",
}

_MIME_TYPE: dict[FileType, str] = {
    FileType.PDF: "application/pdf",
    FileType.CSV: "text/csv",
    # XLSX é convertido para texto antes de enviar — passamos `text/plain`
    # para o client renderizar como `text` block.
    FileType.XLSX: "text/plain",
}

_ALLOWED_EXTENSIONS = {".pdf", ".csv", ".xlsx", ".xls"}


class ParseService:
    """Orquestra a validação + chamada à IA."""

    def __init__(self, anthropic_client: AnthropicClient) -> None:
        self._anthropic = anthropic_client

    async def parse_statement(
        self,
        *,
        file_bytes: bytes,
        filename: str | None,
        max_upload_bytes: int,
    ) -> ExtractedStatement:
        """Valida o upload e chama a Anthropic para extrair movimentações.

        Args:
            file_bytes: bytes do upload (já lido em memória).
            filename: nome original do arquivo. Pode ser `None` (alguns
                clientes HTTP não setam). Usado apenas como heurística da
                extensão e para o `document_kind` do prompt — NUNCA para
                construir caminhos.
            max_upload_bytes: limite efetivo (`Settings.max_upload_bytes`).

        Returns:
            `ExtractedStatement` validado.

        Raises:
            ValidationAppError: arquivo vazio, > limite, extensão proibida,
                magic bytes não reconhecidos, ou `.xls` (não suportado nesta
                versão).
            AnthropicAuthError / AnthropicTimeoutError / AnthropicParseError:
                propagadas do `AnthropicClient` para o handler global.
        """
        self._validate_size(file_bytes, max_upload_bytes)
        self._validate_extension(filename)

        # Magic bytes valida o tipo REAL (rejeita .pdf falsificado contendo XLSX).
        # `_ALLOWED_FOR_PARSE` exclui XLS — mensagem específica abaixo cobre o
        # caso de XLS detectado nos magic bytes.
        from app.utils.magic_bytes import detect_file_type

        detected_raw = detect_file_type(file_bytes)
        if detected_raw == FileType.XLS:
            raise ValidationAppError(
                "Arquivo .xls não é suportado por enquanto.",
                user_message=(
                    "Arquivos .xls (Excel 97-2003) não são suportados. "
                    "Salve como .xlsx ou .csv e tente novamente."
                ),
            )
        detected = validate_upload_type(file_bytes, allowed=_ALLOWED_FOR_PARSE)

        content, mime_type = self._prepare_content(detected, file_bytes)
        document_kind = _DOCUMENT_KIND[detected]

        return await self._anthropic.extract_movements(
            content=content,
            mime_type=mime_type,
            document_kind=document_kind,
        )

    # ------------------------------------------------------------------
    # Validações
    # ------------------------------------------------------------------

    @staticmethod
    def _validate_size(file_bytes: bytes, max_upload_bytes: int) -> None:
        size = len(file_bytes)
        if size == 0:
            raise ValidationAppError(
                "Upload vazio.",
                user_message="O arquivo enviado está vazio.",
            )
        if size > max_upload_bytes:
            raise ValidationAppError(
                f"Upload excede {max_upload_bytes} bytes ({size} recebidos).",
                user_message=(
                    f"O arquivo excede o limite de {max_upload_bytes // (1024 * 1024)} MB."
                ),
            )

    @staticmethod
    def _validate_extension(filename: str | None) -> None:
        """Filtro pela extensão. Magic bytes faz a checagem real depois.

        Aceitar `filename=None` é proposital — alguns clientes HTTP omitem
        o filename e queremos cair no magic bytes em seguida. Quando
        presente, sanitizamos via `PurePosixPath` para evitar path
        traversal em logs (`..\\..\\evil.pdf` vira só `evil.pdf`).
        """
        if filename is None:
            return
        # Sanitização: pega só o nome final, ignora qualquer separador.
        # `PurePosixPath` é seguro mesmo com input vindo do Windows (`\` é
        # tratado como caractere comum, mas o front normaliza; a segunda
        # linha cobre o resto).
        clean = PurePosixPath(filename).name
        clean = clean.replace("\\", "/").rsplit("/", 1)[-1]
        ext = ("." + clean.rsplit(".", 1)[-1].lower()) if "." in clean else ""
        if ext not in _ALLOWED_EXTENSIONS:
            raise ValidationAppError(
                f"Extensão {ext or '(sem extensão)'} não permitida.",
                user_message="Formato não suportado. Envie PDF, CSV, XLS ou XLSX.",
            )

    # ------------------------------------------------------------------
    # Conversão para o formato esperado pela IA
    # ------------------------------------------------------------------

    @staticmethod
    def _prepare_content(file_type: FileType, file_bytes: bytes) -> tuple[bytes, str]:
        """Converte bytes brutos no formato consumido pelo `AnthropicClient`.

        Returns:
            Tupla `(content_bytes, mime_type)`. Para PDF, o `content_bytes`
            é o PDF original — vira `document` block base64. Para outros
            formatos, é o texto utf-8 já extraído — vira `text` block.
        """
        if file_type == FileType.PDF:
            return file_bytes, _MIME_TYPE[FileType.PDF]

        if file_type == FileType.CSV:
            # CSV vai cru (decodificação fica com o client). Magic bytes já
            # garantiu que é texto válido (utf-8 ou latin-1).
            return file_bytes, _MIME_TYPE[FileType.CSV]

        if file_type == FileType.XLSX:
            text = _xlsx_to_text(file_bytes)
            return text.encode("utf-8"), _MIME_TYPE[FileType.XLSX]

        # Qualquer outro tipo deveria ter sido rejeitado por
        # `validate_upload_type` antes — defensivo.
        raise ValidationAppError(
            f"Tipo {file_type} não suportado pelo parsing.",
            user_message="Formato não suportado nesta versão.",
        )


def _xlsx_to_text(file_bytes: bytes) -> str:
    """Renderiza um XLSX como texto tabular (TSV) para a IA.

    Princípio: enviar texto plano é dramaticamente mais barato em tokens do
    que mandar o XLSX cru via document block (que o Claude nem aceita
    nativamente fora de PDFs/imagens). A IA recebe linhas separadas por
    `\\n`, células separadas por `\\t` — formato que o modelo entende
    facilmente.

    Raises:
        ValidationAppError: arquivo XLSX corrompido ou criptografado
            (openpyxl explode com `BadZipFile` / `InvalidFileException`).
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationAppError(
            f"XLSX inválido ou corrompido: {exc}",
            user_message=(
                "Não foi possível ler o arquivo XLSX. Verifique se está íntegro e "
                "sem proteção por senha."
            ),
        ) from exc

    parts: list[str] = []
    for ws in wb.worksheets:
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        parts.append(f"# Aba: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            # Pula linhas totalmente vazias (todas as células vazias após strip).
            if not any(c.strip() for c in cells):
                continue
            parts.append("\t".join(cells))
    wb.close()

    if not parts:
        raise ValidationAppError(
            "XLSX sem dados.",
            user_message="O arquivo XLSX não contém dados legíveis.",
        )
    return "\n".join(parts)

"""Testes do builder de workbook do export Excel (S14 BACK 10.1).

Os testes constroem um `ExportPayload` em memória, geram o XLSX, reabrem
com openpyxl em modo readonly e validam células-chave + cores. Sem DB,
sem rede — função pura.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.modules.reconciliations.export.schemas import (
    AnomalyRow,
    ExportPayload,
    FileEntryRow,
    OmieDivergenceRow,
    SemOmieRow,
    SummarySheetData,
)
from app.modules.reconciliations.export.styles import (
    COLOR_CRITICAL_UNRESOLVED_BG,
    COLOR_DATA_DIVERGENTE,
    COLOR_HEADER_BG,
    COLOR_ROW_ATRASADO,
    COLOR_ROW_CONCILIADO,
    COLOR_ROW_IGNORADO,
    COLOR_ROW_SEM_OMIE,
)
from app.modules.reconciliations.export.workbook import (
    SHEET_NAME_ANOMALIAS,
    SHEET_NAME_DIVERGENCIAS,
    SHEET_NAME_MOVIMENTACAO,
    SHEET_NAME_SEM_OMIE,
    SHEET_NAME_SUMMARY,
    build_workbook,
)

_BRT = timezone(timedelta(hours=-3))


def _summary(**overrides: object) -> SummarySheetData:
    """Factory de SummarySheetData com defaults sensatos."""
    base = SummarySheetData(
        client_name="Cliente Teste S/A",
        bank_name="Sicredi",
        account_name="Sicredi 91263-1",
        period_pt_br="Abril/2026",
        balance_start=Decimal("1000.00"),
        balance_end_file=Decimal("2500.00"),
        balance_end_omie=Decimal("2500.00"),
        balance_difference=Decimal("0.00"),
        total_file_entries=10,
        conciliated_count=7,
        sem_omie_count=2,
        ignored_count=1,
        omie_sem_arquivo_count=3,
        anomaly_total=5,
        anomaly_critical=2,
        anomaly_moderate=2,
        anomaly_info=1,
        anomaly_resolved=3,
        anomaly_critical_unresolved=1,
        generated_at_brt=datetime(2026, 5, 22, 14, 30, tzinfo=_BRT),
        generated_by_email="analista@hologram.com.br",
    )
    if not overrides:
        return base
    # Dataclass frozen → reusa via dataclasses.replace
    from dataclasses import replace

    return replace(base, **overrides)  # type: ignore[arg-type]


def _payload(**overrides: object) -> ExportPayload:
    base = ExportPayload(
        filename="Conciliacao_Cliente_Teste_Sicredi_04-2026",
        is_card=False,
        summary=_summary(),
        file_entries=[],
        omie_divergences=[],
        sem_omie=[],
        anomalies=[],
    )
    if not overrides:
        return base
    from dataclasses import replace

    return replace(base, **overrides)  # type: ignore[arg-type]


def _hex(argb_or_rgb: str | None) -> str | None:
    """Normaliza ARGB→RGB para comparação ('FF1F2937' → '1F2937').

    openpyxl persiste cores como ARGB hex (alpha + RGB) ou como
    `RGB('FF1F2937')`. Comparar substring é confiável e descartável.
    """
    if argb_or_rgb is None:
        return None
    s = str(argb_or_rgb).upper()
    if len(s) == 8:
        return s[-6:]
    return s


# ======================================================================
# Aba 1 — Resumo
# ======================================================================


@pytest.mark.unit
class TestSheet1Summary:
    def test_sheet_exists_with_correct_name(self) -> None:
        buf = build_workbook(_payload())
        wb = load_workbook(buf)
        assert SHEET_NAME_SUMMARY in wb.sheetnames

    def test_header_fields_populated(self) -> None:
        buf = build_workbook(_payload())
        wb = load_workbook(buf)
        ws = wb[SHEET_NAME_SUMMARY]

        # CC (is_card=False) → título bancário tipado (FASE 1 / BACK 1.9).
        assert ws["A1"].value == "CONCILIAÇÃO BANCÁRIA — Sicredi | Sicredi 91263-1 | Abril/2026"
        assert ws["B3"].value == "Cliente Teste S/A"
        assert ws["B4"].value == "Sicredi"
        assert ws["B5"].value == "Sicredi 91263-1"
        assert ws["B6"].value == "Abril/2026"
        assert ws["A8"].value == "Elaborado por Hologram Gestão"

    def test_card_title_is_fatura(self) -> None:
        # Cartão (is_card=True) → título de fatura (FASE 1 / BACK 1.9).
        buf = build_workbook(_payload(is_card=True))
        wb = load_workbook(buf)
        ws = wb[SHEET_NAME_SUMMARY]
        assert ws["A1"].value == "CONCILIAÇÃO DE FATURA — CARTÃO | Sicredi 91263-1 | Abril/2026"

    def test_balance_row_marks_conferido_when_diff_zero(self) -> None:
        buf = build_workbook(_payload())
        wb = load_workbook(buf)
        ws = wb[SHEET_NAME_SUMMARY]
        # Linha 12 = valores; coluna 4 = status
        assert ws.cell(row=12, column=4).value == "Conferido"

    def test_balance_row_marks_divergente_when_diff_positive(self) -> None:
        summary = _summary(balance_difference=Decimal("150.00"))
        buf = build_workbook(_payload(summary=summary))
        wb = load_workbook(buf)
        ws = wb[SHEET_NAME_SUMMARY]
        cell = ws.cell(row=12, column=4)
        assert cell.value == "Divergente"
        # Cor de fundo: vermelho claro (mesmo Atrasado)
        assert _hex(cell.fill.start_color.rgb) == COLOR_ROW_ATRASADO

    def test_indicators_block_lists_all_5(self) -> None:
        buf = build_workbook(_payload())
        wb = load_workbook(buf)
        ws = wb[SHEET_NAME_SUMMARY]
        # Linhas 15..19 = indicadores em ordem fixa.
        assert ws.cell(row=15, column=1).value == "Movimentações no extrato"
        assert ws.cell(row=15, column=2).value == 10
        assert ws.cell(row=16, column=1).value == "Conciliadas"
        assert ws.cell(row=16, column=2).value == 7
        assert ws.cell(row=17, column=1).value == "Sem correspondência Omie"
        assert ws.cell(row=17, column=2).value == 2
        assert ws.cell(row=18, column=1).value == "Omie sem arquivo"
        assert ws.cell(row=18, column=2).value == 3
        assert ws.cell(row=19, column=1).value == "Ignoradas"
        assert ws.cell(row=19, column=2).value == 1

    def test_critical_unresolved_row_is_highlighted_red(self) -> None:
        buf = build_workbook(_payload())
        wb = load_workbook(buf)
        ws = wb[SHEET_NAME_SUMMARY]
        # Bloco de anomalias começa em 21 (15 + 5 indicadores + 1 espaço).
        # Linha "Críticas não resolvidas" = última da lista (1+6).
        # Procura pela label sem depender de offset exato.
        found_row = None
        for row in range(15, 40):
            if ws.cell(row=row, column=1).value == "Críticas não resolvidas":
                found_row = row
                break
        assert found_row is not None, "Label 'Críticas não resolvidas' não encontrada"
        assert ws.cell(row=found_row, column=2).value == 1
        # Fill vermelho (mesma cor para label e valor)
        assert _hex(ws.cell(row=found_row, column=1).fill.start_color.rgb) == (
            COLOR_CRITICAL_UNRESOLVED_BG
        )

    def test_critical_unresolved_not_highlighted_when_zero(self) -> None:
        summary = _summary(anomaly_critical_unresolved=0)
        buf = build_workbook(_payload(summary=summary))
        wb = load_workbook(buf)
        ws = wb[SHEET_NAME_SUMMARY]
        found_row = None
        for row in range(15, 40):
            if ws.cell(row=row, column=1).value == "Críticas não resolvidas":
                found_row = row
                break
        assert found_row is not None
        # Sem destaque vermelho — fill default (None ou "00000000")
        rgb = _hex(ws.cell(row=found_row, column=1).fill.start_color.rgb)
        assert rgb != COLOR_CRITICAL_UNRESOLVED_BG

    def test_footer_contains_email_and_timezone(self) -> None:
        buf = build_workbook(_payload())
        wb = load_workbook(buf)
        ws = wb[SHEET_NAME_SUMMARY]
        # Procura o rodapé varrendo todas as células — rodapé é a última
        # linha utilizada.
        footer_text: str | None = None
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str) and "FaturIA" in value:
                    footer_text = value
                    break
            if footer_text:
                break
        assert footer_text is not None
        assert "analista@hologram.com.br" in footer_text
        assert "(BRT)" in footer_text
        assert "22/05/2026" in footer_text


# ======================================================================
# Aba 2 — Movimentação x Lançamento
# ======================================================================


@pytest.mark.unit
class TestSheet2Movimentacao:
    def _build(self, *rows: FileEntryRow) -> object:
        buf = build_workbook(_payload(file_entries=list(rows)))
        wb = load_workbook(buf)
        return wb[SHEET_NAME_MOVIMENTACAO]

    def test_headers_in_order(self) -> None:
        ws = self._build()
        expected = [
            "Data",
            "Data Omie",
            "Descrição",
            "Valor",
            "Saldo",
            "Fornecedor Omie",
            "Categoria Omie",
            "Análise",
            "Situação",
            "Observação",
        ]
        for col, label in enumerate(expected, start=1):
            assert ws.cell(row=1, column=col).value == label  # type: ignore[attr-defined]

    def test_header_has_header_fill(self) -> None:
        ws = self._build()
        cell = ws.cell(row=1, column=1)  # type: ignore[attr-defined]
        assert _hex(cell.fill.start_color.rgb) == COLOR_HEADER_BG

    def test_conciliado_row_is_green(self) -> None:
        from datetime import date

        row = FileEntryRow(
            transaction_date=date(2026, 4, 10),
            description="Crédito X",
            amount=Decimal("500.00"),
            balance=Decimal("1500.00"),
            supplier="Fornecedor Y",
            category="Receita",
            situation="conciliado",
            user_note=None,
        )
        ws = self._build(row)
        # Linha 2 = primeira data row
        for col in range(1, 9):
            cell = ws.cell(row=2, column=col)  # type: ignore[attr-defined]
            assert _hex(cell.fill.start_color.rgb) == COLOR_ROW_CONCILIADO, f"col {col}"

    def test_sem_omie_row_is_yellow(self) -> None:
        from datetime import date

        row = FileEntryRow(
            transaction_date=date(2026, 4, 11),
            description="Débito sem match",
            amount=Decimal("-300.00"),
            balance=None,
            supplier=None,
            category=None,
            situation="sem_omie",
            user_note=None,
        )
        ws = self._build(row)
        cell = ws.cell(row=2, column=1)  # type: ignore[attr-defined]
        assert _hex(cell.fill.start_color.rgb) == COLOR_ROW_SEM_OMIE

    def test_ignorado_row_is_gray(self) -> None:
        from datetime import date

        row = FileEntryRow(
            transaction_date=date(2026, 4, 12),
            description="Ignorada",
            amount=Decimal("-10.00"),
            balance=None,
            supplier=None,
            category=None,
            situation="ignorado",
            user_note=None,
        )
        ws = self._build(row)
        cell = ws.cell(row=2, column=1)  # type: ignore[attr-defined]
        assert _hex(cell.fill.start_color.rgb) == COLOR_ROW_IGNORADO

    def test_decimal_value_persisted_as_number(self) -> None:
        """Valor monetário grava como número (Decimal) — não string."""
        from datetime import date

        row = FileEntryRow(
            transaction_date=date(2026, 4, 10),
            description="Teste",
            amount=Decimal("-1234.56"),
            balance=Decimal("100.00"),
            supplier=None,
            category=None,
            situation="conciliado",
            user_note=None,
        )
        ws = self._build(row)
        amount_cell = ws.cell(row=2, column=4)  # type: ignore[attr-defined]
        # openpyxl retorna Decimal/float — nunca string
        assert not isinstance(amount_cell.value, str)
        assert amount_cell.value == Decimal("-1234.56") or amount_cell.value == -1234.56

    def test_placeholder_for_missing_supplier(self) -> None:
        from datetime import date

        row = FileEntryRow(
            transaction_date=date(2026, 4, 10),
            description="Sem cache",
            amount=Decimal("10.00"),
            balance=None,
            supplier=None,
            category=None,
            situation="conciliado",
            user_note=None,
        )
        ws = self._build(row)
        assert ws.cell(row=2, column=6).value == "—"  # type: ignore[attr-defined]
        assert ws.cell(row=2, column=7).value == "—"  # type: ignore[attr-defined]


# ======================================================================
# Aba 2 — coluna "Data Omie" (cartão: FASE 1 / BACK 1.9)
# ======================================================================


@pytest.mark.unit
class TestSheet2MovimentacaoCartao:
    def _build_card(self, *rows: FileEntryRow) -> object:
        buf = build_workbook(_payload(is_card=True, file_entries=list(rows)))
        wb = load_workbook(buf)
        return wb[SHEET_NAME_MOVIMENTACAO]

    def test_data_omie_column_after_data(self) -> None:
        ws = self._build_card()
        assert ws.cell(row=1, column=1).value == "Data"  # type: ignore[attr-defined]
        assert ws.cell(row=1, column=2).value == "Data Omie"  # type: ignore[attr-defined]
        assert ws.cell(row=1, column=3).value == "Descrição"  # type: ignore[attr-defined]

    def test_divergente_fills_data_omie_orange(self) -> None:
        from datetime import date

        row = FileEntryRow(
            transaction_date=date(2026, 4, 10),
            description="Compra divergente",
            amount=Decimal("-250.00"),
            balance=None,
            supplier="Loja X",
            category="Compras",
            situation="conciliado_data_divergente",
            user_note=None,
            omie_date=date(2026, 4, 12),
        )
        ws = self._build_card(row)
        omie_cell = ws.cell(row=2, column=2)  # type: ignore[attr-defined]
        val = omie_cell.value
        assert val is not None
        assert (getattr(val, "year", 0), getattr(val, "month", 0), getattr(val, "day", 0)) == (
            2026,
            4,
            12,
        )
        assert _hex(omie_cell.fill.start_color.rgb) == COLOR_DATA_DIVERGENTE

    def test_conciliado_sem_divergencia_data_omie_vazia(self) -> None:
        from datetime import date

        row = FileEntryRow(
            transaction_date=date(2026, 4, 10),
            description="Compra exata",
            amount=Decimal("-100.00"),
            balance=None,
            supplier="Loja Y",
            category="Compras",
            situation="conciliado",
            user_note=None,
            omie_date=date(2026, 4, 10),
        )
        ws = self._build_card(row)
        # Conciliado sem divergência → "Data Omie" vazia (mesmo tendo match).
        assert ws.cell(row=2, column=2).value is None  # type: ignore[attr-defined]


# ======================================================================
# Aba 2 — "Data Omie" em CONTA CORRENTE
# ======================================================================
# A coluna nasceu só no cartão, mas a FASE 1 passou a aplicar
# `DATE_DIVERGENCE_RANGE` também à CC (CLAUDE.md §5.2): existem linhas
# `conciliado_data_divergente` em conta corrente, e o relatório não dizia
# divergente de QUÊ — a data do Omie ficava só no payload. O buraco não apareceu
# antes porque a cobertura da coluna existia apenas no caminho de cartão.


@pytest.mark.unit
class TestSheet2MovimentacaoContaCorrente:
    def _build_cc(self, *rows: FileEntryRow) -> object:
        buf = build_workbook(_payload(is_card=False, file_entries=list(rows)))
        wb = load_workbook(buf)
        return wb[SHEET_NAME_MOVIMENTACAO]

    def _row(self, situation: str, omie_date: date | None) -> FileEntryRow:
        return FileEntryRow(
            transaction_date=date(2026, 7, 7),
            description="Boleto fornecedor",
            amount=Decimal("-890.00"),
            balance=None,
            supplier="Fornecedor Z",
            category="Despesa",
            situation=situation,
            user_note=None,
            omie_date=omie_date,
        )

    def test_coluna_existe_mesmo_sem_nenhuma_divergencia(self) -> None:
        # Layout estável entre meses: a coluna não aparece e some conforme o mês
        # tenha ou não divergência — quem compara exports lado a lado depende disso.
        ws = self._build_cc(self._row("conciliado", date(2026, 7, 7)))
        assert ws.cell(row=1, column=2).value == "Data Omie"  # type: ignore[attr-defined]
        assert ws.cell(row=1, column=3).value == "Descrição"  # type: ignore[attr-defined]

    def test_linha_divergente_traz_a_data_do_omie_destacada(self) -> None:
        ws = self._build_cc(self._row("conciliado_data_divergente", date(2026, 7, 10)))
        cell = ws.cell(row=2, column=2)  # type: ignore[attr-defined]
        val = cell.value
        assert (getattr(val, "year", 0), getattr(val, "month", 0), getattr(val, "day", 0)) == (
            2026,
            7,
            10,
        )
        # Laranja sobrescreve o verde de conciliado só nessa célula.
        assert _hex(cell.fill.start_color.rgb) == COLOR_DATA_DIVERGENTE

    def test_conciliado_exato_deixa_a_celula_vazia(self) -> None:
        ws = self._build_cc(self._row("conciliado", date(2026, 7, 7)))
        assert ws.cell(row=2, column=2).value is None  # type: ignore[attr-defined]

    def test_sem_omie_deixa_a_celula_vazia(self) -> None:
        ws = self._build_cc(self._row("sem_omie", None))
        assert ws.cell(row=2, column=2).value is None  # type: ignore[attr-defined]

    def test_descricao_e_valor_seguem_legiveis_apos_o_deslocamento(self) -> None:
        # A coluna nova empurra tudo uma casa à direita; este teste falha se
        # alguém reintroduzir o layout antigo em algum ponto do caminho.
        ws = self._build_cc(self._row("conciliado", None))
        assert ws.cell(row=2, column=3).value == "Boleto fornecedor"  # type: ignore[attr-defined]
        assert ws.cell(row=2, column=4).value == Decimal("-890.00")  # type: ignore[attr-defined]


# ======================================================================
# Aba 3 — Divergências Omie
# ======================================================================


@pytest.mark.unit
class TestSheet3Divergencias:
    def _build(self, *rows: OmieDivergenceRow) -> object:
        buf = build_workbook(_payload(omie_divergences=list(rows)))
        wb = load_workbook(buf)
        return wb[SHEET_NAME_DIVERGENCIAS]

    def test_atrasado_row_is_red(self) -> None:
        from datetime import date

        row = OmieDivergenceRow(
            transaction_date=date(2026, 4, 5),
            supplier="Fornecedor X",
            category="Aluguel",
            amount=Decimal("-1500.00"),
            omie_status="Atrasado",
            user_note=None,
        )
        ws = self._build(row)
        for col in range(1, 7):
            cell = ws.cell(row=2, column=col)  # type: ignore[attr-defined]
            assert _hex(cell.fill.start_color.rgb) == COLOR_ROW_ATRASADO, f"col {col}"

    def test_previsto_row_has_no_color(self) -> None:
        from datetime import date

        row = OmieDivergenceRow(
            transaction_date=date(2026, 4, 5),
            supplier="X",
            category="Y",
            amount=Decimal("100.00"),
            omie_status="Previsto",
            user_note=None,
        )
        ws = self._build(row)
        cell = ws.cell(row=2, column=1)  # type: ignore[attr-defined]
        # Sem fill aplicado
        assert _hex(cell.fill.start_color.rgb) != COLOR_ROW_ATRASADO

    def test_missing_amount_shows_placeholder(self) -> None:
        from datetime import date

        row = OmieDivergenceRow(
            transaction_date=date(2026, 4, 5),
            supplier=None,
            category=None,
            amount=None,
            omie_status="Previsto",
            user_note=None,
        )
        ws = self._build(row)
        # Coluna 4 = Valor
        assert ws.cell(row=2, column=4).value == "—"  # type: ignore[attr-defined]


# ======================================================================
# Aba 4 — Sem Omie
# ======================================================================


@pytest.mark.unit
class TestSheet4SemOmie:
    def _build(self, *rows: SemOmieRow) -> object:
        buf = build_workbook(_payload(sem_omie=list(rows)))
        wb = load_workbook(buf)
        return wb[SHEET_NAME_SEM_OMIE]

    def test_headers_match_spec(self) -> None:
        ws = self._build()
        assert ws.cell(row=1, column=1).value == "Data"  # type: ignore[attr-defined]
        assert ws.cell(row=1, column=2).value == "Descrição"  # type: ignore[attr-defined]
        assert ws.cell(row=1, column=3).value == "Valor"  # type: ignore[attr-defined]
        assert ws.cell(row=1, column=4).value == "Observação"  # type: ignore[attr-defined]

    def test_all_rows_yellow(self) -> None:
        from datetime import date

        rows = [
            SemOmieRow(
                transaction_date=date(2026, 4, 5),
                description="A",
                amount=Decimal("-10.00"),
                user_note=None,
            ),
            SemOmieRow(
                transaction_date=date(2026, 4, 6),
                description="B",
                amount=Decimal("-20.00"),
                user_note="Nota",
            ),
        ]
        ws = self._build(*rows)
        for row_idx in (2, 3):
            for col in range(1, 5):
                cell = ws.cell(row=row_idx, column=col)  # type: ignore[attr-defined]
                assert _hex(cell.fill.start_color.rgb) == COLOR_ROW_SEM_OMIE


# ======================================================================
# Aba 5 — Anomalias
# ======================================================================


@pytest.mark.unit
class TestSheet5Anomalias:
    def _build(self, *rows: AnomalyRow) -> object:
        buf = build_workbook(_payload(anomalies=list(rows)))
        wb = load_workbook(buf)
        return wb[SHEET_NAME_ANOMALIAS]

    def test_coluna_detalhe_traz_o_motivo_da_anomalia(self) -> None:
        """O report da Bruna (04/08/2026): o relatório vinha "pela metade".

        Sem esta coluna, várias linhas do mesmo tipo ficam indistinguíveis entre
        si no Excel, embora a tela mostre o porquê de cada uma. Era o campo
        `context`, que o export nunca decifrava.
        """
        rows = [
            AnomalyRow(
                context="Extrato indica Getulio da Silva Rios mas fornecedor Omie é Renilson",
                severity="critical",
                type_name="Qualificação incoerente",
                related_line="07/07/2026 · Pix Getulio · -R$ 2.800,00",
                detected_by="ai",
                resolved=False,
                resolution_note=None,
            ),
            AnomalyRow(
                context="Data arquivo: 07/07/2026 · Data Omie: 10/07/2026",
                severity="moderate",
                type_name="Data divergente",
                related_line="07/07/2026 · Pix Cleidson · -R$ 1.400,00",
                detected_by="ai",
                resolved=False,
                resolution_note=None,
            ),
        ]
        ws = self._build(*rows)

        assert ws.cell(row=1, column=3).value == "Detalhe"  # type: ignore[attr-defined]
        assert (  # type: ignore[attr-defined]
            ws.cell(row=2, column=3).value
            == "Extrato indica Getulio da Silva Rios mas fornecedor Omie é Renilson"
        )
        # Na linha de data divergente, o que sumia era justamente QUAL é a data
        # do Omie — o dado que torna a anomalia acionável.
        assert (  # type: ignore[attr-defined]
            ws.cell(row=3, column=3).value == "Data arquivo: 07/07/2026 · Data Omie: 10/07/2026"
        )

    def test_anomalia_estrutural_sem_contexto_sai_com_celula_vazia(self) -> None:
        """Ausência legítima de contexto não é falha — estruturais não têm."""
        rows = [
            AnomalyRow(
                context="",
                severity="critical",
                type_name="Movimentação sem lançamento no Omie",
                related_line="07/07/2026 · Pix Maiane · -R$ 2.800,00",
                detected_by="ai",
                resolved=False,
                resolution_note=None,
            )
        ]
        ws = self._build(*rows)

        # `None` e não `""`: o openpyxl relê célula vazia como None. O que
        # importa é não haver texto, e sobretudo não haver "[indecifrável]" —
        # ausência de contexto não é falha de decifragem.
        assert not ws.cell(row=2, column=3).value  # type: ignore[attr-defined]

    def test_linha_relacionada_traz_a_descricao_da_movimentacao(self) -> None:
        """A tela sempre mostrou a descrição; o Excel trazia só data e valor."""
        rows = [
            AnomalyRow(
                context="",
                severity="info",
                type_name="X",
                related_line="10/07/2026 · Pix enviado: Cp:18236120-Cleidson · -R$ 2.800,00",
                detected_by="ai",
                resolved=False,
                resolution_note=None,
            )
        ]
        ws = self._build(*rows)

        assert ws.cell(row=1, column=4).value == "Linha relacionada"  # type: ignore[attr-defined]
        assert (  # type: ignore[attr-defined]
            ws.cell(row=2, column=4).value
            == "10/07/2026 · Pix enviado: Cp:18236120-Cleidson · -R$ 2.800,00"
        )

    def test_preserva_a_ordem_recebida_sem_reordenar(self) -> None:
        """A aba 5 escreve na ordem que recebe — quem ordena é o serviço.

        Até 07/08/2026 o workbook reordenava por (severidade, resolvida) na
        montagem, e era isso que fazia a planilha contar uma história diferente
        da tela. A ordem canônica agora é cronológica e vive em
        `anomaly_ordering`, consultada pela tela e pelo export. Se alguém voltar
        a ordenar aqui, este teste quebra.
        """
        rows = [
            AnomalyRow(
                context="",
                severity="info",
                type_name="1º — informativa",
                related_line="01/07/2026 · R$ 100,00",
                detected_by="ai",
                resolved=True,
                resolution_note="ok",
            ),
            AnomalyRow(
                context="",
                severity="critical",
                type_name="2º — crítica",
                related_line="05/07/2026 · R$ 200,00",
                detected_by="ai",
                resolved=False,
                resolution_note=None,
            ),
            AnomalyRow(
                context="",
                severity="moderate",
                type_name="3º — moderada",
                related_line="—",
                detected_by="manual",
                resolved=False,
                resolution_note=None,
            ),
        ]
        ws = self._build(*rows)

        # Nem a severidade nem o status de resolvida movem a linha de lugar.
        assert ws.cell(row=2, column=2).value == "1º — informativa"  # type: ignore[attr-defined]
        assert ws.cell(row=3, column=2).value == "2º — crítica"  # type: ignore[attr-defined]
        assert ws.cell(row=4, column=2).value == "3º — moderada"  # type: ignore[attr-defined]

    def test_resolvida_nao_e_mais_empurrada_para_o_fim(self) -> None:
        """Mudança de comportamento deliberada, junto com a ordem cronológica.

        A planilha empurrava as resolvidas para o fim de cada faixa de
        severidade. A tela nunca fez isso — ela oferece o filtro de status. Como
        o critério agora é o mesmo nos dois lugares, a resolvida fica na posição
        cronológica dela, e quem quiser separar usa o filtro.
        """
        rows = [
            AnomalyRow(
                context="",
                severity="critical",
                type_name="C1 — RESOLVIDA",
                related_line="01/07/2026 · R$ 100,00",
                detected_by="ai",
                resolved=True,
                resolution_note="ok",
            ),
            AnomalyRow(
                context="",
                severity="critical",
                type_name="C2 — pendente",
                related_line="02/07/2026 · R$ 200,00",
                detected_by="ai",
                resolved=False,
                resolution_note=None,
            ),
        ]
        ws = self._build(*rows)

        assert ws.cell(row=2, column=2).value == "C1 — RESOLVIDA"  # type: ignore[attr-defined]
        assert ws.cell(row=3, column=2).value == "C2 — pendente"  # type: ignore[attr-defined]

    def test_detected_by_translates_to_pt(self) -> None:
        rows = [
            AnomalyRow(
                context="",
                severity="info",
                type_name="X",
                related_line="—",
                detected_by="ai",
                resolved=False,
                resolution_note=None,
            ),
            AnomalyRow(
                context="",
                severity="info",
                type_name="Y",
                related_line="—",
                detected_by="manual",
                resolved=False,
                resolution_note=None,
            ),
        ]
        ws = self._build(*rows)
        # Coluna 5: a aba ganhou "Detalhe" na 3 e tudo depois dela deslocou.
        # `ai` agora sai como "Sistema", igual à tela — nem toda detecção
        # automática usa modelo (as estruturais são código determinístico).
        seen = {ws.cell(row=2, column=5).value, ws.cell(row=3, column=5).value}  # type: ignore[attr-defined]
        assert seen == {"Sistema", "Manual"}

    def test_status_label_pt(self) -> None:
        rows = [
            AnomalyRow(
                context="",
                severity="info",
                type_name="X",
                related_line="—",
                detected_by="ai",
                resolved=True,
                resolution_note="feito",
            ),
            AnomalyRow(
                context="",
                severity="info",
                type_name="Y",
                related_line="—",
                detected_by="ai",
                resolved=False,
                resolution_note=None,
            ),
        ]
        ws = self._build(*rows)
        # Na ordem recebida — o rótulo é o que está sob teste, não a posição.
        assert ws.cell(row=2, column=6).value == "Resolvida"  # type: ignore[attr-defined]
        assert ws.cell(row=3, column=6).value == "Pendente"  # type: ignore[attr-defined]

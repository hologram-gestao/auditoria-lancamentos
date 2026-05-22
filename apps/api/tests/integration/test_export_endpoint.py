"""Testes de integração do endpoint de export Excel (S14 BACK 10.1).

Cobre:
    - 401 sem auth.
    - 200 admin baixa Excel de qualquer cliente (binário parseável).
    - 200 manager baixa Excel da própria carteira.
    - 404 manager fora da carteira.
    - 404 sessão inexistente.
    - 404 sessão soft-deletada.
    - 409 sessão em status processing.
    - 409 sessão em status error.
    - Excel contém abas + dados criptografados em claro.
    - Filename do Content-Disposition é sanitizado.
"""

from __future__ import annotations

import hashlib
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.crypto import encrypt
from app.core.security import hash_password
from app.db.models import (
    AnomalySeverity,
    AnomalyType,
    Client,
    ClientAssignment,
    OmieAccountCache,
    ReconciliationAnomaly,
    ReconciliationFileEntry,
    ReconciliationSession,
    User,
    UserRole,
)

if TYPE_CHECKING:
    from httpx import AsyncClient


ADMIN_EMAIL = "export-admin@hologram.com.br"
MANAGER_A_EMAIL = "export-mgr-a@hologram.com.br"
MANAGER_B_EMAIL = "export-mgr-b@hologram.com.br"
PLAIN_PASSWORD = "Senh@ForteParaTeste#1"

FAKE_APP_KEY = "test-app-key-export"
FAKE_APP_SECRET = "test-app-secret-export"


def _hex64(salt: str) -> str:
    return hashlib.sha256(salt.encode()).hexdigest()


async def _seed_user(session: AsyncSession, *, email: str, role: UserRole) -> User:
    user = User(
        name="T",
        email=email.lower(),
        password_hash=hash_password(PLAIN_PASSWORD),
        role=role.value,
        active=True,
    )
    session.add(user)
    await session.flush()
    return user


async def _seed_client(
    session: AsyncSession,
    *,
    creator: User,
    manager: User | None = None,
    name: str = "Cliente Export S/A",
) -> Client:
    hex_key = get_settings().OMIE_ENCRYPTION_KEY.get_secret_value()
    ct_k, iv_k = encrypt(FAKE_APP_KEY, hex_key)
    ct_s, iv_s = encrypt(FAKE_APP_SECRET, hex_key)
    cli = Client(
        name=name,
        omie_app_key_encrypted=ct_k,
        omie_app_key_iv=iv_k,
        omie_app_secret_encrypted=ct_s,
        omie_app_secret_iv=iv_s,
        active=True,
        created_by=creator.id,
    )
    session.add(cli)
    await session.flush()
    if manager is not None:
        session.add(ClientAssignment(client_id=cli.id, user_id=manager.id, assigned_by=creator.id))
        await session.flush()
    return cli


async def _seed_account_cache(
    session: AsyncSession, *, client: Client, omie_conta_id: int
) -> OmieAccountCache:
    cache_row = OmieAccountCache(
        client_id=client.id,
        omie_conta_id=omie_conta_id,
        name="Sicredi 91263-1",
        bank_name="Sicredi",
        account_type="CC",
    )
    session.add(cache_row)
    await session.flush()
    return cache_row


async def _seed_session(
    session: AsyncSession,
    *,
    client: Client,
    creator: User,
    status: str = "reviewing",
    deleted_at: datetime | None = None,
    omie_conta_id: int = 42,
) -> ReconciliationSession:
    sess = ReconciliationSession(
        client_id=client.id,
        created_by=creator.id,
        omie_conta_id=omie_conta_id,
        reference_month=date(2026, 4, 1),
        date_tolerance_days=3,
        file_hash=_hex64(f"export-{uuid4().hex}"),
        status=status,
        balance_start=Decimal("1000.00"),
        balance_end_file=Decimal("1500.00"),
        balance_end_omie=Decimal("1500.00"),
        balance_difference=Decimal("0.00"),
        processed_at=datetime.now(UTC),
        total_file_entries=2,
        conciliated_count=0,
        sem_omie_count=2,
        omie_sem_arquivo_count=0,
        anomaly_count=1,
        deleted_at=deleted_at,
    )
    session.add(sess)
    await session.flush()
    return sess


async def _seed_file_entry(
    session: AsyncSession,
    *,
    recon: ReconciliationSession,
    description: str,
    amount: Decimal,
    situation: str = "sem_omie",
    tx_date: date = date(2026, 4, 10),
) -> ReconciliationFileEntry:
    hex_key = get_settings().OMIE_ENCRYPTION_KEY.get_secret_value()
    ct, iv = encrypt(description, hex_key)
    entry = ReconciliationFileEntry(
        session_id=recon.id,
        transaction_date=tx_date,
        description_encrypted=ct,
        description_iv=iv,
        amount=amount,
        situation=situation,
    )
    session.add(entry)
    await session.flush()
    return entry


async def _seed_anomaly(
    session: AsyncSession,
    *,
    recon: ReconciliationSession,
    severity: str = AnomalySeverity.CRITICAL.value,
    resolved: bool = False,
) -> ReconciliationAnomaly:
    atype = AnomalyType(
        code=f"export-{uuid4().hex[:8]}",
        name="Anomalia Teste Crítica",
        description="x",
        severity=severity,
        active=True,
    )
    session.add(atype)
    await session.flush()
    anomaly = ReconciliationAnomaly(
        session_id=recon.id,
        anomaly_type_id=atype.id,
        detected_by="ai",
        resolved=resolved,
    )
    session.add(anomaly)
    await session.flush()
    return anomaly


async def _login(client: AsyncClient, email: str) -> None:
    resp = await client.post(
        "/api/v1/auth/login",
        json={"email": email, "password": PLAIN_PASSWORD},
    )
    assert resp.status_code == 200, resp.text


# ----------------------------------------------------------------------
# Tests
# ----------------------------------------------------------------------


@pytest.mark.integration
class TestExportEndpoint:
    async def test_unauthenticated_returns_401(self, client_with_db: AsyncClient) -> None:
        sid = uuid4()
        resp = await client_with_db.post(f"/api/v1/reconciliations/{sid}/export")
        assert resp.status_code == 401

    async def test_admin_downloads_xlsx_with_correct_headers(
        self, client_with_db: AsyncClient, db_session: AsyncSession
    ) -> None:
        admin = await _seed_user(db_session, email=ADMIN_EMAIL, role=UserRole.ADMIN)
        cli = await _seed_client(db_session, creator=admin)
        await _seed_account_cache(db_session, client=cli, omie_conta_id=42)
        sess = await _seed_session(db_session, client=cli, creator=admin)
        await _seed_file_entry(
            db_session,
            recon=sess,
            description="Pagamento Padaria",
            amount=Decimal("-150.00"),
        )
        await _seed_file_entry(
            db_session,
            recon=sess,
            description="Recebimento Cielo",
            amount=Decimal("250.00"),
        )
        await _seed_anomaly(db_session, recon=sess)
        await _login(client_with_db, ADMIN_EMAIL)

        resp = await client_with_db.post(f"/api/v1/reconciliations/{sess.id}/export")
        assert resp.status_code == 200, resp.text

        # Content-Type correto
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Content-Disposition: attachment + filename sanitizado
        cd = resp.headers["content-disposition"]
        assert cd.startswith("attachment")
        assert ".xlsx" in cd
        assert "Conciliacao_" in cd
        # NomeCliente sanitizado (sem barra) e MesAno 04-2026
        assert "04-2026" in cd

        # Binário válido: openpyxl consegue parsear
        buf = BytesIO(resp.content)
        wb = load_workbook(buf)
        assert "Resumo" in wb.sheetnames
        assert "Movimentação x Lançamento" in wb.sheetnames
        assert "Divergências Omie" in wb.sheetnames
        assert "Sem Omie" in wb.sheetnames
        assert "Anomalias" in wb.sheetnames

        # Aba Resumo: cliente + período
        summary = wb["Resumo"]
        assert summary["B3"].value == "Cliente Export S/A"
        assert summary["B6"].value == "Abril/2026"

        # Aba 2: descrições descriptografadas
        mov = wb["Movimentação x Lançamento"]
        descriptions = {mov.cell(row=r, column=2).value for r in (2, 3)}
        assert descriptions == {"Pagamento Padaria", "Recebimento Cielo"}

    async def test_manager_in_portfolio_downloads_xlsx(
        self, client_with_db: AsyncClient, db_session: AsyncSession
    ) -> None:
        admin = await _seed_user(db_session, email=ADMIN_EMAIL, role=UserRole.ADMIN)
        mgr = await _seed_user(db_session, email=MANAGER_A_EMAIL, role=UserRole.MANAGER)
        cli = await _seed_client(db_session, creator=admin, manager=mgr)
        await _seed_account_cache(db_session, client=cli, omie_conta_id=42)
        sess = await _seed_session(db_session, client=cli, creator=admin)
        await _seed_file_entry(db_session, recon=sess, description="X", amount=Decimal("-10.00"))
        await _login(client_with_db, MANAGER_A_EMAIL)

        resp = await client_with_db.post(f"/api/v1/reconciliations/{sess.id}/export")
        assert resp.status_code == 200, resp.text

    async def test_manager_outside_portfolio_returns_404(
        self, client_with_db: AsyncClient, db_session: AsyncSession
    ) -> None:
        admin = await _seed_user(db_session, email=ADMIN_EMAIL, role=UserRole.ADMIN)
        mgr_a = await _seed_user(db_session, email=MANAGER_A_EMAIL, role=UserRole.MANAGER)
        await _seed_user(db_session, email=MANAGER_B_EMAIL, role=UserRole.MANAGER)
        cli = await _seed_client(db_session, creator=admin, manager=mgr_a)
        sess = await _seed_session(db_session, client=cli, creator=admin)
        await _login(client_with_db, MANAGER_B_EMAIL)

        resp = await client_with_db.post(f"/api/v1/reconciliations/{sess.id}/export")
        assert resp.status_code == 404

    async def test_nonexistent_session_returns_404(
        self, client_with_db: AsyncClient, db_session: AsyncSession
    ) -> None:
        await _seed_user(db_session, email=ADMIN_EMAIL, role=UserRole.ADMIN)
        await _login(client_with_db, ADMIN_EMAIL)
        resp = await client_with_db.post(f"/api/v1/reconciliations/{uuid4()}/export")
        assert resp.status_code == 404

    async def test_soft_deleted_session_returns_404(
        self, client_with_db: AsyncClient, db_session: AsyncSession
    ) -> None:
        admin = await _seed_user(db_session, email=ADMIN_EMAIL, role=UserRole.ADMIN)
        cli = await _seed_client(db_session, creator=admin)
        sess = await _seed_session(
            db_session,
            client=cli,
            creator=admin,
            deleted_at=datetime.now(UTC),
        )
        await _login(client_with_db, ADMIN_EMAIL)

        resp = await client_with_db.post(f"/api/v1/reconciliations/{sess.id}/export")
        assert resp.status_code == 404

    @pytest.mark.parametrize("blocked_status", ["processing", "error"])
    async def test_non_exportable_status_returns_409(
        self,
        client_with_db: AsyncClient,
        db_session: AsyncSession,
        blocked_status: str,
    ) -> None:
        admin = await _seed_user(db_session, email=ADMIN_EMAIL, role=UserRole.ADMIN)
        cli = await _seed_client(db_session, creator=admin)
        sess = await _seed_session(db_session, client=cli, creator=admin, status=blocked_status)
        await _login(client_with_db, ADMIN_EMAIL)

        resp = await client_with_db.post(f"/api/v1/reconciliations/{sess.id}/export")
        assert resp.status_code == 409, resp.text
        body = resp.json()
        assert body["error"]["code"] == "CONFLICT"
        assert "exportada" in body["error"]["userMessage"]

    async def test_done_status_is_exportable(
        self, client_with_db: AsyncClient, db_session: AsyncSession
    ) -> None:
        admin = await _seed_user(db_session, email=ADMIN_EMAIL, role=UserRole.ADMIN)
        cli = await _seed_client(db_session, creator=admin)
        sess = await _seed_session(db_session, client=cli, creator=admin, status="done")
        await _seed_file_entry(db_session, recon=sess, description="X", amount=Decimal("-1.00"))
        await _login(client_with_db, ADMIN_EMAIL)

        resp = await client_with_db.post(f"/api/v1/reconciliations/{sess.id}/export")
        assert resp.status_code == 200, resp.text

    async def test_filename_sanitizes_special_chars(
        self, client_with_db: AsyncClient, db_session: AsyncSession
    ) -> None:
        admin = await _seed_user(db_session, email=ADMIN_EMAIL, role=UserRole.ADMIN)
        cli = await _seed_client(db_session, creator=admin, name="Padaria São João / Filial *")
        await _seed_account_cache(db_session, client=cli, omie_conta_id=42)
        sess = await _seed_session(db_session, client=cli, creator=admin)
        await _seed_file_entry(db_session, recon=sess, description="X", amount=Decimal("-1.00"))
        await _login(client_with_db, ADMIN_EMAIL)

        resp = await client_with_db.post(f"/api/v1/reconciliations/{sess.id}/export")
        assert resp.status_code == 200, resp.text

        cd = resp.headers["content-disposition"]
        # Sem caracteres inválidos no filename ASCII
        for forbidden in r"\/:*?<>|":
            # `filename="..."` ASCII part — extrai entre primeiras aspas.
            ascii_part = cd.split('filename="', 1)[1].split('"', 1)[0]
            assert forbidden not in ascii_part, f"'{forbidden}' encontrado em {ascii_part}"
        # Acento removido
        assert "Sao_Joao" in cd or "Sao Joao" in cd
